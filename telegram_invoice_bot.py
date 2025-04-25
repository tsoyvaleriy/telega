import os
import re
import PyPDF2
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
from datetime import datetime
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

invoices_data = []

def extract_data_from_pdf(file_path):
    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        text = "\n".join([page.extract_text() or "" for page in reader.pages])
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    name, address, model, vin = "", "", "", ""

    for line in lines:
        if "vin code" in line.lower() or "vincode" in line.lower():
            match = re.search(r'([A-HJ-NPR-Z0-9]{17})', line)
            if match:
                vin = match.group(1)
                model_text = line.split(vin)[0].strip()
                model_text = model_text.replace("FULL DESCRIPTION OF GOODS", "").strip()
                model = model_text.split(")")[0] + ")" if ")" in model_text else model_text
                break

    for line in lines:
        l = line.upper()
        if sum(k in l for k in ["UL.", "GOR.", "OBL.", "RUSSIA", "KV.", "D.", "KRAY", "G."]) >= 2:
            if "RUSSIA" in line.upper():
                r_idx = line.upper().find("RUSSIA") + len("RUSSIA")
                address = line[:r_idx].strip()
                break
            break

    try:
        images = convert_from_path(file_path, dpi=300)
        text = pytesseract.image_to_string(images[-1], lang="eng+rus")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        for i, line in enumerate(lines):
            if "АДРЕС" in line.upper() and i >= 1:
                name = lines[i - 1]
                break
    except Exception as e:
        print("OCR error:", e)

    return {"NAME BL": name, "ADDRESS": address, "MODEL+YEAR": model, "VIN": vin}

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Отправь PDF-инвойсы. Когда закончишь — напиши /done")

async def handle_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = update.message.document
    if file.mime_type != 'application/pdf':
        await update.message.reply_text("Пожалуйста, отправь PDF-файл.")
        return
    file_path = f"{file.file_id}.pdf"
    new_file = await context.bot.get_file(file.file_id)
    await new_file.download_to_drive(file_path)
    try:
        data = extract_data_from_pdf(file_path)
        invoices_data.append(data)
        await update.message.reply_text("Инвойс добавлен ✅")
    except Exception as e:
        await update.message.reply_text(f"Ошибка при обработке: {str(e)}")
    finally:
        os.remove(file_path)

async def finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not invoices_data:
        await update.message.reply_text("Нет данных для экспорта.")
        return
    wb = load_workbook("образец.xlsx")
    ws = wb.active
    for i, row_data in enumerate(invoices_data):
        row = 2 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=row_data.get("NAME BL", ""))
        ws.cell(row=row, column=3, value=row_data.get("ADDRESS", ""))
        ws.cell(row=row, column=4, value=row_data.get("MODEL+YEAR", ""))
        ws.cell(row=row, column=5, value=row_data.get("VIN", ""))
    filename = f"Отправки Владивосток {datetime.now().strftime('%d.%m.%Y')}.xlsx"
    wb.save(filename)
    await update.message.reply_document(document=open(filename, "rb"), filename=filename)
    os.remove(filename)
    invoices_data.clear()

if __name__ == '__main__':
    TOKEN = "7690456905:AAEKZJxv0ofO0BE1MiiXGzfLCI67magIvRg"
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("done", finish))
    app.add_handler(MessageHandler(filters.Document.PDF, handle_pdf))
    print("Бот запущен ✅")
    app.run_polling()
